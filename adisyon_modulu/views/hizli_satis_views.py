# views/hizli_satis_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Q, Sum
from datetime import datetime, timedelta
import json
import qrcode
from io import BytesIO
import base64
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import io

from ..models import (
    Sube, HizliSatisCihaz, HizliSatisKasa, HizliSatisUrun,
    HizliSatisSepet, HizliSatisSepetItem, HizliSatis,
    HizliSatisKasaHareket, Musteri, HizliSatisItem
)
from ..printing import yaziciya_veri_gonder

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def _hizli_satis_kasa_idleri_kullanici(user):
    """Kullanıcının görebileceği aktif kasa id’leri (şube kısıtlı veya tümü)."""
    qs = HizliSatisKasa.objects.filter(aktif=True)
    if hasattr(user, 'profil') and getattr(user.profil, 'sube_id', None):
        qs = qs.filter(sube_id=user.profil.sube_id)
    return set(qs.values_list('id', flat=True))


def get_or_create_active_sepet(kasa, kullanici):
    """
    Aynı kasa için tek aktif sepet garantiler.
    Fazla aktif sepet varsa en sonuncuyu bırakır, diğerlerini pasif yapar.
    """
    aktif_sepetler = HizliSatisSepet.objects.filter(
        kasa=kasa,
        aktif=True
    ).order_by("-id")

    sepet = aktif_sepetler.first()

    if sepet:
        aktif_sepetler.exclude(id=sepet.id).update(aktif=False)
        if sepet.kullanici_id != kullanici.id:
            sepet.kullanici = kullanici
            sepet.save(update_fields=["kullanici"])
        return sepet

    return HizliSatisSepet.objects.create(
        kasa=kasa,
        kullanici=kullanici,
        aktif=True
    )


@login_required
def kasa_secim(request):
    """Kullanıcının yetkili olduğu şubedeki kasaları listele"""
    if hasattr(request.user, 'profil') and request.user.profil.sube:
        kasalar = HizliSatisKasa.objects.filter(
            sube=request.user.profil.sube,
            aktif=True,
        ).order_by('liste_sirasi', 'kasa_no')
        return render(request, 'hizli_satis/kasa_secim.html', {
            'kasalar': kasalar,
            'sube': request.user.profil.sube
        })

    kasalar = HizliSatisKasa.objects.filter(aktif=True).order_by('liste_sirasi', 'kasa_no')
    return render(request, 'hizli_satis/kasa_secim.html', {'kasalar': kasalar})


@login_required
@require_POST
def kasa_liste_sira_kaydet(request):
    """Hesap (kasa) kartlarının sürükle-bırak sırasını kaydeder."""
    allowed = _hizli_satis_kasa_idleri_kullanici(request.user)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        order = payload.get('order')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': 'Geçersiz veri'}, status=400)
    if not isinstance(order, list) or not order:
        return JsonResponse({'success': False, 'error': 'Sıra listesi gerekli'}, status=400)
    try:
        ids = [int(x) for x in order]
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Geçersiz ID'}, status=400)
    if not set(ids).issubset(allowed) or len(ids) != len(set(ids)):
        return JsonResponse({'success': False, 'error': 'Yetkisiz veya geçersiz kayıt'}, status=403)
    for idx, kid in enumerate(ids):
        HizliSatisKasa.objects.filter(id=kid).update(liste_sirasi=(idx + 1) * 10)
    return JsonResponse({'success': True})


@login_required
@require_POST
def kasa_favori_bir_kaydet(request):
    """Tek şubede varsayılan (Favori 1) kasayı ayarlar."""
    allowed = _hizli_satis_kasa_idleri_kullanici(request.user)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        kasa_id = payload.get('kasa_id')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': 'Geçersiz veri'}, status=400)
    try:
        kasa_id = int(kasa_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Kasa ID gerekli'}, status=400)
    if kasa_id not in allowed:
        return JsonResponse({'success': False, 'error': 'Yetkisiz'}, status=403)
    kasa = get_object_or_404(HizliSatisKasa, id=kasa_id, aktif=True)
    kasa.favori_bir = True
    kasa.save()
    request.session['son_kasa_id'] = kasa_id
    return JsonResponse({'success': True})


@login_required
def kasa_ekrani(request, kasa_id):
    """Hızlı satış ana ekranı"""
    kasa = get_object_or_404(HizliSatisKasa, id=kasa_id, aktif=True)
    request.session['son_kasa_id'] = kasa_id

    sepet = get_or_create_active_sepet(kasa, request.user)

    urunler = HizliSatisUrun.objects.filter(
    sube=kasa.sube,
    aktif=True,
    stok_miktari__gt=0
    ).select_related("urun", "sube").order_by('urun__ad')

    context = {
        'kasa': kasa,
        'sepet': sepet,
        'urunler': urunler,
        'toplam_urun': urunler.count(),
    }
    return render(request, 'hizli_satis/kasa_ekrani.html', context)


@login_required
@csrf_exempt
def barkod_oku(request):
    """Barkod/karekod okuma API'si"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST gerekli'})

    try:
        data = json.loads(request.body)
        barkod = data.get('barkod')
        kasa_id = data.get('kasa_id')

        if not barkod or not kasa_id:
            return JsonResponse({'success': False, 'error': 'Barkod ve kasa ID gerekli'})

        kasa = get_object_or_404(HizliSatisKasa, id=kasa_id, aktif=True)

        urun = get_object_or_404(
            HizliSatisUrun,
            barkod=barkod,
            sube=kasa.sube,
            aktif=True
        )

        sepet = get_or_create_active_sepet(kasa, request.user)

        mevcut_item = sepet.sepet_items.filter(urun=urun, aktif=True).first()
        guncel_fiyat = Decimal(str(urun.gecerli_fiyat())).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        if mevcut_item:
            yeni_adet = mevcut_item.adet + 1
            if urun.stok_miktari < yeni_adet:
                return JsonResponse({
                    'success': False,
                    'error': f'Yetersiz stok! Mevcut: {urun.stok_miktari}'
                })

            mevcut_item.adet = yeni_adet
            mevcut_item.birim_fiyat = guncel_fiyat
            mevcut_item.toplam_fiyat = (guncel_fiyat * Decimal(str(mevcut_item.adet))).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP
            )
            mevcut_item.save(update_fields=['adet', 'birim_fiyat', 'toplam_fiyat'])
            item = mevcut_item
            mesaj = f"{urun.urun.ad} adedi artırıldı"
        else:
            if urun.stok_miktari < 1:
                return JsonResponse({
                    'success': False,
                    'error': f'Yetersiz stok! Mevcut: {urun.stok_miktari}'
                })

            item = HizliSatisSepetItem.objects.create(
                sepet=sepet,
                urun=urun,
                adet=1,
                birim_fiyat=guncel_fiyat,
                toplam_fiyat=guncel_fiyat
            )
            mesaj = f"{urun.urun.ad} sepete eklendi"

        sepet.sepet_guncelle()
        sepet.refresh_from_db()

        items = []
        for i in sepet.sepet_items.filter(aktif=True):
            items.append({
                'id': i.id,
                'urun_adi': i.urun.urun.ad,
                'adet': float(i.adet),
                'birim_fiyat': float(i.birim_fiyat),
                'toplam_fiyat': float(i.toplam_fiyat),
                'indirim_yuzde': float(i.indirim_yuzde)
            })

        return JsonResponse({
            'success': True,
            'mesaj': mesaj,
            'urun': {
                'id': urun.id,
                'ad': urun.urun.ad,
                'barkod': urun.barkod,
                'fiyat': float(urun.gecerli_fiyat()),
                'stok': float(urun.stok_miktari),
                'indirimde': urun.indirimde_mi,
                'indirimli_fiyat': float(urun.indirimli_fiyat) if urun.indirimli_fiyat else None,
            },
            'sepet_item': {
                'id': item.id,
                'adet': float(item.adet),
                'toplam_fiyat': float(item.toplam_fiyat)
            },
            'items': items,
            'sepet_detay': {
                'ara_toplam': float(sepet.ara_toplam),
                'indirim_tutari': float(sepet.indirim_tutari),
                'kdv_tutari': float(sepet.kdv_tutari),
                'genel_toplam': float(sepet.genel_toplam)
            }
        })

    except HizliSatisUrun.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Ürün bulunamadı'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@csrf_exempt
def sepet_guncelle(request, sepet_id):
    """Sepetteki ürünleri güncelle (adet değiştir, sil)"""
    sepet = get_object_or_404(HizliSatisSepet, id=sepet_id, kullanici=request.user, aktif=True)

    if request.method == 'POST':
        data = json.loads(request.body)
        item_id = data.get('item_id')
        action = data.get('action')

        item = get_object_or_404(HizliSatisSepetItem, id=item_id, sepet=sepet, aktif=True)
        guncel_fiyat = Decimal(str(item.urun.gecerli_fiyat())).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        if action == 'artir':
            if item.urun.stok_miktari < item.adet + 1:
                return JsonResponse({
                    'success': False,
                    'error': f'Yetersiz stok! Mevcut: {item.urun.stok_miktari}'
                })
            item.adet += 1
            item.birim_fiyat = guncel_fiyat
            item.toplam_fiyat = (guncel_fiyat * Decimal(str(item.adet))).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP
            )
            item.save(update_fields=['adet', 'birim_fiyat', 'toplam_fiyat'])
            mesaj = "Adet artırıldı"

        elif action == 'azalt':
            if item.adet > 1:
                item.adet -= 1
                item.birim_fiyat = guncel_fiyat
                item.toplam_fiyat = (guncel_fiyat * Decimal(str(item.adet))).quantize(
                    Decimal('0.01'),
                    rounding=ROUND_HALF_UP
                )
                item.save(update_fields=['adet', 'birim_fiyat', 'toplam_fiyat'])
                mesaj = "Adet azaltıldı"
            else:
                item.aktif = False
                item.save(update_fields=['aktif'])
                mesaj = "Ürün sepetten çıkarıldı"

        elif action == 'sil':
            item.aktif = False
            item.save(update_fields=['aktif'])
            mesaj = "Ürün sepetten çıkarıldı"

        else:
            return JsonResponse({'success': False, 'error': 'Geçersiz işlem'})

        sepet.sepet_guncelle()

        if not sepet.sepet_items.filter(aktif=True).exists():
            sepet.aktif = False
            sepet.save(update_fields=['aktif'])

        items = []
        for i in sepet.sepet_items.filter(aktif=True):
            items.append({
                'id': i.id,
                'urun_adi': i.urun.urun.ad,
                'adet': float(i.adet),
                'birim_fiyat': float(i.birim_fiyat),
                'toplam_fiyat': float(i.toplam_fiyat),
                'indirim_yuzde': float(i.indirim_yuzde)
            })

        return JsonResponse({
            'success': True,
            'mesaj': mesaj,
            'items': items,
            'sepet_detay': {
                'ara_toplam': float(sepet.ara_toplam),
                'indirim_tutari': float(sepet.indirim_tutari),
                'kdv_tutari': float(sepet.kdv_tutari),
                'genel_toplam': float(sepet.genel_toplam)
            }
        })

    return JsonResponse({'success': False, 'error': 'POST gerekli'})


@login_required
def sepet_temizle(request, sepet_id):
    """Sepeti tamamen temizle"""
    sepet = get_object_or_404(HizliSatisSepet, id=sepet_id, kullanici=request.user, aktif=True)

    if request.method == 'POST':
        sepet.sepet_items.filter(aktif=True).update(aktif=False)
        sepet.aktif = False
        sepet.save(update_fields=['aktif'])

        yeni_sepet = get_or_create_active_sepet(sepet.kasa, request.user)

        messages.success(request, "Sepet temizlendi. Yeni sepet oluşturuldu.")

        return JsonResponse({'success': True, 'yeni_sepet_id': yeni_sepet.id})

    return JsonResponse({'success': False, 'error': 'Geçersiz istek'})


@login_required
def odeme_ekrani(request, sepet_id):
    """Ödeme ekranını göster"""
    sepet = get_object_or_404(HizliSatisSepet, id=sepet_id, kullanici=request.user, aktif=True)

    if not sepet.sepet_items.filter(aktif=True).exists():
        messages.warning(request, "Sepetiniz boş!")
        return redirect('kasa_ekrani', kasa_id=sepet.kasa.id)

    son_musteriler = Musteri.objects.order_by('-son_ziyaret')[:10]

    context = {
        'sepet': sepet,
        'son_musteriler': son_musteriler,
        'odeme_tipleri': HizliSatis.ODEME_TIPLERI,
    }
    return render(request, 'hizli_satis/odeme_ekrani.html', context)


@login_required
@csrf_exempt
def odeme_tamamla(request, sepet_id):
    """Ödemeyi tamamla ve satışı kaydet"""
    sepet = get_object_or_404(HizliSatisSepet, id=sepet_id, kullanici=request.user, aktif=True)

    if request.method == 'POST':
        data = json.loads(request.body)
        odeme_tipi = data.get('odeme_tipi', 'nakit')

        if 'tutar' in data:
            tutar = Decimal(str(data.get('tutar', 0)))

            if odeme_tipi == 'nakit':
                nakit = tutar
                kart = Decimal('0')
            else:
                nakit = Decimal('0')
                kart = tutar
        else:
            nakit = Decimal(str(data.get('nakit', 0)))
            kart = Decimal(str(data.get('kart', 0)))

        musteri_id = data.get('musteri_id')

        toplam_odeme = nakit + kart
        if toplam_odeme < sepet.genel_toplam:
            return JsonResponse({
                'success': False,
                'error': f'Eksik ödeme! Toplam: {sepet.genel_toplam:.2f} TL, Ödenen: {toplam_odeme:.2f} TL'
            })

        para_ustu = toplam_odeme - sepet.genel_toplam

        musteri = None
        if musteri_id:
            musteri = get_object_or_404(Musteri, id=musteri_id)

        fis_no = f"HS-{sepet.kasa.sube.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}-{sepet.id}"

        satis = HizliSatis.objects.create(
            kasa=sepet.kasa,
            sube=sepet.kasa.sube,
            musteri=musteri,
            fis_no=fis_no,
            toplam_tutar=sepet.genel_toplam,
            odeme_tipi=odeme_tipi,
            nakit_odenen=nakit,
            kart_odenen=kart,
            para_ustu=para_ustu,
            kullanici=request.user
        )

        for item in sepet.sepet_items.filter(aktif=True):
            HizliSatisItem.objects.create(
                hizli_satis=satis,
                urun=item.urun.urun,
                adet=item.adet,
                birim_fiyat=item.birim_fiyat,
                toplam_fiyat=item.toplam_fiyat,
                stok_dusuldu=True
            )

            item.urun.stok_miktari -= item.adet
            item.urun.save(update_fields=['stok_miktari'])

        kasa = sepet.kasa
        kasa.gunluk_ciro += satis.toplam_tutar
        kasa.gunluk_satis_sayisi += 1
        kasa.son_islem_tarihi = timezone.now()
        kasa.son_islem_tutari = satis.toplam_tutar
        kasa.save()

        sepet.aktif = False
        sepet.save(update_fields=['aktif'])

        yazdir_basarili = fis_yazdir(satis)

        yeni_sepet = get_or_create_active_sepet(kasa, request.user)

        return JsonResponse({
            'success': True,
            'satis_id': satis.id,
            'fis_no': fis_no,
            'para_ustu': float(para_ustu),
            'yazdirildi': yazdir_basarili,
            'yeni_sepet_id': yeni_sepet.id,
            'fis_data': {
                'sube_adi': satis.sube.ad,
                'kasa_adi': satis.kasa.kasa_adi,
                'fis_no': fis_no,
                'tarih': satis.tarih.strftime('%d.%m.%Y %H:%M'),
                'items': [
                    {
                        'urun_adi': item.urun.ad,
                        'adet': float(item.adet),
                        'birim_fiyat': float(item.birim_fiyat),
                        'toplam_fiyat': float(item.toplam_fiyat)
                    } for item in satis.items.all()
                ],
                'ara_toplam': float(sepet.ara_toplam),
                'indirim_tutari': float(sepet.indirim_tutari),
                'kdv_tutari': float(sepet.kdv_tutari),
                'genel_toplam': float(satis.toplam_tutar),
                'odeme_tipi': satis.get_odeme_tipi_display(),
                'nakit': float(nakit),
                'kart': float(kart),
                'para_ustu': float(para_ustu)
            }
        })

    return JsonResponse({'success': False, 'error': 'POST gerekli'})


def fis_yazdir(satis):
    """Satış fişini yazdır"""
    try:
        kasa = satis.kasa
        if not kasa or not kasa.yazici:
            return False

        yazici = kasa.yazici

        def turkce_duzelt(metin):
            donusum = {
                'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G',
                'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S',
                'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C',
            }
            for turkce, ascii_ in donusum.items():
                metin = metin.replace(turkce, ascii_)
            return metin

        ESC = b'\x1b'
        GS = b'\x1d'

        komutlar = bytearray()
        komutlar.extend(ESC + b'\x40')

        komutlar.extend(ESC + b'\x61' + b'\x01')
        komutlar.extend(ESC + b'\x21' + b'\x30')
        komutlar.extend("LAR".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(ESC + b'\x21' + b'\x00')
        komutlar.extend("HIZLI SATIS".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(turkce_duzelt(f"Sube: {satis.sube.ad}").encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"Fis No: {satis.fis_no}".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"Tarih: {satis.tarih.strftime('%d.%m.%Y %H:%M')}".encode('utf-8'))
        komutlar.extend(b'\x0a\x0a')

        if satis.musteri:
            komutlar.extend(f"Musteri: {turkce_duzelt(satis.musteri.ad_soyad)}".encode('utf-8'))
            komutlar.extend(b'\x0a\x0a')

        komutlar.extend(ESC + b'\x61' + b'\x00')
        komutlar.extend(("=" * 32).encode('utf-8'))
        komutlar.extend(b'\x0a')

        for item in satis.items.all():
            urun_adi = turkce_duzelt(item.urun.ad)
            komutlar.extend(f"{urun_adi}".encode('utf-8'))
            komutlar.extend(b'\x0a')
            komutlar.extend(f"  {float(item.adet)} x {float(item.birim_fiyat):.2f}TL".encode('utf-8'))
            komutlar.extend(b'\x0a')
            komutlar.extend(ESC + b'\x21' + b'\x11')
            komutlar.extend(f"  = {float(item.toplam_fiyat):.2f}TL".encode('utf-8'))
            komutlar.extend(b'\x0a')
            komutlar.extend(ESC + b'\x21' + b'\x00')
            komutlar.extend(("-" * 32).encode('utf-8'))
            komutlar.extend(b'\x0a')

        komutlar.extend(("=" * 32).encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"TOPLAM: {float(satis.toplam_tutar):.2f}TL".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"ODEME: {satis.get_odeme_tipi_display()}".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"NAKIT: {float(satis.nakit_odenen):.2f}TL".encode('utf-8'))
        komutlar.extend(b'\x0a')
        komutlar.extend(f"KART: {float(satis.kart_odenen):.2f}TL".encode('utf-8'))
        komutlar.extend(b'\x0a')

        if satis.para_ustu > 0:
            komutlar.extend(f"PARA USTU: {float(satis.para_ustu):.2f}TL".encode('utf-8'))
            komutlar.extend(b'\x0a')

        komutlar.extend(b'\x0a')
        komutlar.extend("TESEKKUR EDERIZ".encode('utf-8'))
        komutlar.extend(b'\x0a\x0a')

        komutlar.extend(GS + b'\x56' + b'\x41' + b'\x00')

        yaziciya_veri_gonder(yazici, komutlar)

        return True

    except Exception as e:
        print(f"Fiş yazdırma hatası: {e}")
        return False


@login_required
def karekod_olustur(request, urun_id):
    """Ürün için karekod oluştur"""
    urun = get_object_or_404(HizliSatisUrun, id=urun_id)

    qr_data = json.dumps({
        'barkod': urun.barkod,
        'urun': urun.urun.ad,
        'fiyat': float(urun.satis_fiyati),
        'sube': urun.sube.id
    })

    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    buffer = BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()

    return JsonResponse({
        'success': True,
        'karekod': f'data:image/png;base64,{img_str}'
    })


@login_required
def gun_sonu(request, kasa_id):
    """Kasa için gün sonu raporu"""
    kasa = get_object_or_404(HizliSatisKasa, id=kasa_id)

    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                data = json.loads(request.body)
                note = data.get('note', '')
            except Exception:
                note = request.POST.get('note', '')
        else:
            note = request.POST.get('note', '')

        bugun = timezone.now().date()
        satislar = HizliSatis.objects.filter(
            kasa=kasa,
            tarih__date=bugun
        )

        nakit_toplam = satislar.filter(odeme_tipi='nakit').aggregate(
            toplam=Sum('toplam_tutar')
        )['toplam'] or 0

        kart_toplam = satislar.filter(odeme_tipi='kart').aggregate(
            toplam=Sum('toplam_tutar')
        )['toplam'] or 0

        karma_satislar = satislar.filter(odeme_tipi='karma')
        for satis in karma_satislar:
            nakit_toplam += satis.nakit_odenen or 0
            kart_toplam += satis.kart_odenen or 0

        toplam_ciro = nakit_toplam + kart_toplam
        satis_sayisi = satislar.count()

        urun_raporu = HizliSatisItem.objects.filter(
            hizli_satis__in=satislar
        ).values(
            'urun__ad'
        ).annotate(
            adet_toplam=Sum('adet'),
            kazanc_toplam=Sum('toplam_fiyat')
        ).order_by('-kazanc_toplam')

        saatlik_veriler = {}
        max_ciro = 0

        for saat in range(24):
            saatlik_satislar = satislar.filter(tarih__hour=saat)
            saat_ciro = saatlik_satislar.aggregate(
                toplam=Sum('toplam_tutar')
            )['toplam'] or 0

            saatlik_veriler[f"{saat:02d}"] = {
                'satis': saatlik_satislar.count(),
                'ciro': float(saat_ciro)
            }

            if saat_ciro > max_ciro:
                max_ciro = float(saat_ciro)

        HizliSatisKasaHareket.objects.create(
            kasa=kasa,
            hareket_tipi='gun_sonu',
            tutar=toplam_ciro,
            aciklama=f"Gün sonu - Satış: {satis_sayisi}, Ciro: {toplam_ciro:.2f} TL. Not: {note}",
            kullanici=request.user
        )

        kasa.gunluk_ciro = 0
        kasa.gunluk_satis_sayisi = 0
        kasa.save()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'toplam_ciro': float(toplam_ciro),
                'satis_sayisi': satis_sayisi,
                'nakit_toplam': float(nakit_toplam),
                'kart_toplam': float(kart_toplam),
                'mesaj': f"Gün sonu tamamlandı! Toplam ciro: {toplam_ciro:.2f} TL"
            })

        messages.success(request, f"Gün sonu tamamlandı! Toplam ciro: {toplam_ciro:.2f} TL")
        return redirect('kasa_secim')

    bugun = timezone.now().date()
    satislar = HizliSatis.objects.filter(
        kasa=kasa,
        tarih__date=bugun
    ).order_by('-tarih')

    nakit_toplam = satislar.filter(odeme_tipi='nakit').aggregate(
        toplam=Sum('toplam_tutar')
    )['toplam'] or 0

    kart_toplam = satislar.filter(odeme_tipi='kart').aggregate(
        toplam=Sum('toplam_tutar')
    )['toplam'] or 0

    karma_satislar = satislar.filter(odeme_tipi='karma')
    for satis in karma_satislar:
        nakit_toplam += satis.nakit_odenen or 0
        kart_toplam += satis.kart_odenen or 0

    toplam_ciro = nakit_toplam + kart_toplam

    urun_raporu = HizliSatisItem.objects.filter(
        hizli_satis__in=satislar
    ).values(
        'urun__ad'
    ).annotate(
        adet_toplam=Sum('adet'),
        kazanc_toplam=Sum('toplam_fiyat')
    ).order_by('-kazanc_toplam')[:10]

    saatlik_veriler = {}
    max_ciro = 0

    for saat in range(24):
        saatlik_satislar = satislar.filter(tarih__hour=saat)
        saat_ciro = saatlik_satislar.aggregate(
            toplam=Sum('toplam_tutar')
        )['toplam'] or 0

        saatlik_veriler[f"{saat:02d}"] = {
            'satis': saatlik_satislar.count(),
            'ciro': float(saat_ciro)
        }

        if saat_ciro > max_ciro:
            max_ciro = float(saat_ciro)

    context = {
        'kasa': kasa,
        'satislar': satislar,
        'toplam_ciro': toplam_ciro,
        'nakit_toplam': nakit_toplam,
        'kart_toplam': kart_toplam,
        'satis_sayisi': satislar.count(),
        'urun_raporu': urun_raporu,
        'saatlik_veriler': saatlik_veriler,
        'max_ciro': max_ciro,
        'bugun': bugun,
    }
    return render(request, 'hizli_satis/gun_sonu.html', context)


@login_required
def satis_gecmisi(request, kasa_id):
    """Kasanın satış geçmişi"""
    kasa = get_object_or_404(HizliSatisKasa, id=kasa_id)

    request.session['son_kasa_id'] = kasa_id

    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    musteri_adi = request.GET.get('musteri')

    satislar = HizliSatis.objects.filter(kasa=kasa)

    if baslangic:
        satislar = satislar.filter(tarih__date__gte=baslangic)
    if bitis:
        satislar = satislar.filter(tarih__date__lte=bitis)
    if musteri_adi:
        satislar = satislar.filter(musteri__ad_soyad__icontains=musteri_adi)

    satislar = satislar.order_by('-tarih')[:100]

    toplam_tutar = satislar.aggregate(toplam=Sum('toplam_tutar'))['toplam'] or 0

    context = {
        'kasa': kasa,
        'satislar': satislar,
        'toplam_tutar': toplam_tutar,
        'baslangic': baslangic,
        'bitis': bitis,
    }
    return render(request, 'hizli_satis/satis_gecmisi.html', context)


@login_required
def satis_export(request, format_type):
    """Satış geçmişini Excel veya PDF olarak dışa aktar"""
    kasa_id = request.GET.get('kasa_id')

    if not kasa_id and hasattr(request.user, 'profil') and request.user.profil.sube:
        kasa_id = request.session.get('son_kasa_id')

    if not kasa_id and hasattr(request.user, 'profil') and getattr(request.user.profil, 'sube_id', None):
        fav = HizliSatisKasa.objects.filter(
            sube_id=request.user.profil.sube_id,
            aktif=True,
            favori_bir=True,
        ).first()
        if fav:
            kasa_id = str(fav.id)

    if not kasa_id:
        fav = HizliSatisKasa.objects.filter(aktif=True, favori_bir=True).first()
        if fav:
            kasa_id = str(fav.id)

    if not kasa_id:
        return JsonResponse({'error': 'Kasa ID gerekli'}, status=400)

    kasa = get_object_or_404(HizliSatisKasa, id=kasa_id)

    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    musteri_adi = request.GET.get('musteri')
    odeme_tipi = request.GET.get('odeme')

    satislar = HizliSatis.objects.filter(kasa=kasa).select_related('musteri').prefetch_related('items__urun')

    if baslangic:
        satislar = satislar.filter(tarih__date__gte=baslangic)
    if bitis:
        satislar = satislar.filter(tarih__date__lte=bitis)
    if musteri_adi:
        satislar = satislar.filter(musteri__ad_soyad__icontains=musteri_adi)
    if odeme_tipi:
        satislar = satislar.filter(odeme_tipi=odeme_tipi)

    satislar = satislar.order_by('-tarih')

    if format_type == 'excel':
        return export_to_excel(satislar, kasa)
    if format_type == 'pdf':
        return export_to_pdf(satislar, kasa, request)

    return HttpResponse("Geçersiz format", status=400)


def export_to_excel(satislar, kasa):
    """Satışları Excel dosyasına aktar"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satış Geçmişi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"{kasa.sube.ad} - {kasa.kasa_adi} Satış Geçmişi Raporu"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    date_cell = ws['A2']
    date_cell.value = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal="center")

    headers = ['Fiş No', 'Tarih', 'Saat', 'Müşteri', 'Ürün Sayısı', 'Ödeme Tipi', 'Tutar (TL)', 'Detay']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_num = 5
    toplam_tutar = 0

    for satis in satislar:
        urun_detay = []
        for item in satis.items.all():
            urun_detay.append(f"{item.urun.ad} x{item.adet} = {item.toplam_fiyat:.2f}TL")

        if satis.odeme_tipi == 'nakit':
            odeme_tipi = "Nakit"
        elif satis.odeme_tipi == 'kart':
            odeme_tipi = "Kart"
        else:
            odeme_tipi = "Karma"

        ws.cell(row=row_num, column=1, value=satis.fis_no).border = thin_border
        ws.cell(row=row_num, column=2, value=satis.tarih.strftime('%d.%m.%Y')).border = thin_border
        ws.cell(row=row_num, column=3, value=satis.tarih.strftime('%H:%M:%S')).border = thin_border
        ws.cell(row=row_num, column=4, value=satis.musteri.ad_soyad if satis.musteri else "-").border = thin_border
        ws.cell(row=row_num, column=5, value=satis.items.count()).border = thin_border
        ws.cell(row=row_num, column=6, value=odeme_tipi).border = thin_border
        ws.cell(row=row_num, column=7, value=float(satis.toplam_tutar)).border = thin_border
        ws.cell(row=row_num, column=8, value="\n".join(urun_detay)).border = thin_border

        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="right")

        toplam_tutar += satis.toplam_tutar
        row_num += 1

    row_num += 1
    toplam_cell = ws.cell(row=row_num, column=6, value="GENEL TOPLAM:")
    toplam_cell.font = Font(bold=True)
    toplam_cell.alignment = Alignment(horizontal="right")
    toplam_cell.border = thin_border

    toplam_tutar_cell = ws.cell(row=row_num, column=7, value=float(toplam_tutar))
    toplam_tutar_cell.font = Font(bold=True, color="FF0000")
    toplam_tutar_cell.alignment = Alignment(horizontal="right")
    toplam_tutar_cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="satis_gecmisi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


def export_to_pdf(satislar, kasa, request):
    """Satışları PDF dosyasına aktar"""
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm
    )

    elements = []
    styles = getSampleStyleSheet()

    try:
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
    except Exception:
        pass

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=20
    )

    title = Paragraph(f"{kasa.sube.ad} - {kasa.kasa_adi} Satış Geçmişi Raporu", title_style)
    elements.append(title)

    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        textColor=colors.gray,
        spaceAfter=30
    )
    date_text = Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style)
    elements.append(date_text)

    headers = ['Fiş No', 'Tarih', 'Saat', 'Müşteri', 'Ürün', 'Ödeme', 'Tutar (TL)']
    data = [headers]
    toplam_tutar = 0

    for satis in satislar:
        if satis.odeme_tipi == 'nakit':
            odeme = "Nakit"
        elif satis.odeme_tipi == 'kart':
            odeme = "Kart"
        else:
            odeme = "Karma"

        musteri_adi = satis.musteri.ad_soyad if satis.musteri else "-"
        if len(musteri_adi) > 20:
            musteri_adi = musteri_adi[:20] + "..."

        row = [
            satis.fis_no,
            satis.tarih.strftime('%d.%m.%Y'),
            satis.tarih.strftime('%H:%M'),
            musteri_adi,
            str(satis.items.count()),
            odeme,
            f"{satis.toplam_tutar:.2f}"
        ]
        data.append(row)
        toplam_tutar += satis.toplam_tutar

    data.append(['', '', '', '', '', 'GENEL TOPLAM:', f"{toplam_tutar:.2f}"])

    table = Table(data, colWidths=[3 * cm, 2.5 * cm, 2 * cm, 4 * cm, 2 * cm, 2.5 * cm, 3 * cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
    ]))

    elements.append(table)

    stats_style = ParagraphStyle(
        'StatsStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceBefore=20,
        spaceAfter=10
    )

    stats_text = f"""
    <b>Rapor Özeti:</b><br/>
    Toplam Satış: {satislar.count()} adet<br/>
    Toplam Ciro: {toplam_tutar:.2f} TL<br/>
    Ortalama Sepet: {(toplam_tutar / satislar.count()) if satislar.count() > 0 else 0:.2f} TL
    """

    elements.append(Paragraph(stats_text, stats_style))

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="satis_gecmisi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    return response


@login_required
def api_musteri_ara(request):
    """Müşteri arama API'si"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse([], safe=False)

    musteriler = Musteri.objects.filter(
        Q(ad_soyad__icontains=query) | Q(telefon__icontains=query)
    )[:10]

    data = [{
        'id': m.id,
        'ad_soyad': m.ad_soyad,
        'telefon': m.telefon,
        'sadakat_seviyesi': m.sadakat_seviyesi
    } for m in musteriler]

    return JsonResponse(data, safe=False)


@login_required
def api_satis_detay(request, satis_id):
    """Satış detayı API'si"""
    satis = get_object_or_404(HizliSatis, id=satis_id)

    urunler = []
    for item in satis.items.all():
        urunler.append({
            'ad': item.urun.ad,
            'adet': float(item.adet),
            'birim_fiyat': float(item.birim_fiyat),
            'toplam': float(item.toplam_fiyat)
        })

    data = {
        'id': satis.id,
        'fis_no': satis.fis_no,
        'tarih': satis.tarih.strftime('%d.%m.%Y %H:%M:%S'),
        'kasa': satis.kasa.kasa_adi,
        'sube': satis.sube.ad,
        'musteri': {
            'id': satis.musteri.id if satis.musteri else None,
            'ad': satis.musteri.ad_soyad if satis.musteri else None,
            'telefon': satis.musteri.telefon if satis.musteri else None
        } if satis.musteri else None,
        'urunler': urunler,
        'ara_toplam': float(sum(item.toplam_fiyat for item in satis.items.all())),
        'indirim': 0,
        'toplam': float(satis.toplam_tutar),
        'odeme_tipi': satis.get_odeme_tipi_display(),
        'nakit': float(satis.nakit_odenen or 0),
        'kart': float(satis.kart_odenen or 0),
        'para_ustu': float(satis.para_ustu or 0)
    }

    return JsonResponse(data)


@login_required
def fis_yazdir_sayfa(request, satis_id):
    """Fiş yazdırma sayfası"""
    satis = get_object_or_404(HizliSatis, id=satis_id)

    context = {
        'satis': satis,
        'items': satis.items.all(),
    }
    return render(request, 'hizli_satis/fis_yazdir.html', context)
