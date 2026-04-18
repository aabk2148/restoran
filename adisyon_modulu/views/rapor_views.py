from datetime import datetime, timedelta
from io import BytesIO

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db import connection
from django.db.models import Avg, Count, DecimalField, ExpressionWrapper, F, Q, Sum, fields
from django.db.utils import ProgrammingError
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..models import Adisyon, Gider, IptalKaydi, SiparisItem, Sube
from .auth_views import format_tl, rapor_gorebilir_mi


def _db_column_exists(model, column_name):
    table_name = model._meta.db_table
    with connection.cursor() as cursor:
        columns = connection.introspection.get_table_description(cursor, table_name)
    return any(column.name == column_name for column in columns)


def _date_range(request):
    bas_str = request.GET.get('baslangic')
    bit_str = request.GET.get('bitis')
    bugun = timezone.localdate()

    try:
        baslangic = datetime.strptime(bas_str, '%Y-%m-%d').date() if bas_str else bugun
        bitis = datetime.strptime(bit_str, '%Y-%m-%d').date() if bit_str else bugun
    except ValueError:
        baslangic = bitis = bugun

    if baslangic > bitis:
        baslangic, bitis = bitis, baslangic
    return baslangic, bitis, bugun


def _build_report_context(request, sube_id=None):
    baslangic, bitis, bugun = _date_range(request)
    secili_sube = None
    secili_sube_id = sube_id or request.GET.get('sube')

    adisyonlar = Adisyon.objects.filter(
        durum='Kapali',
        acilis_zamani__date__gte=baslangic,
        acilis_zamani__date__lte=bitis,
    ).select_related('sube', 'garson', 'masa')

    iptallar = IptalKaydi.objects.filter(
        zaman__date__gte=baslangic,
        zaman__date__lte=bitis,
    ).select_related('garson', 'sube').order_by('-zaman')

    giderler = Gider.objects.filter(
        tarih__gte=baslangic,
        tarih__lte=bitis,
    ).select_related('sube').order_by('-tarih')

    if secili_sube_id:
        secili_sube = get_object_or_404(Sube, id=secili_sube_id)
        adisyonlar = adisyonlar.filter(sube=secili_sube)
        iptallar = iptallar.filter(sube=secili_sube)
        giderler = giderler.filter(sube=secili_sube)

    subeler = Sube.objects.all().order_by('ad')

    sube_ozetleri = []
    if not secili_sube:
        sube_ozetleri = list(
            Adisyon.objects.filter(
                durum='Kapali',
                acilis_zamani__date__gte=baslangic,
                acilis_zamani__date__lte=bitis,
            )
            .values('sube_id', 'sube__ad')
            .annotate(
                adisyon_sayisi=Count('id'),
                nakit=Coalesce(Sum('nakit_odenen'), 0, output_field=DecimalField()),
                kart=Coalesce(Sum('kart_odenen'), 0, output_field=DecimalField()),
            )
            .order_by('-nakit', '-kart')
        )
        for item in sube_ozetleri:
            item['id'] = item['sube_id']
            item['ad'] = item['sube__ad']
            item['ciro'] = item['nakit'] + item['kart']
        sube_ozetleri = sorted(sube_ozetleri, key=lambda row: row['ciro'], reverse=True)

    nakit = adisyonlar.aggregate(t=Coalesce(Sum('nakit_odenen'), 0, output_field=DecimalField()))['t']
    kart = adisyonlar.aggregate(t=Coalesce(Sum('kart_odenen'), 0, output_field=DecimalField()))['t']
    indirim = adisyonlar.aggregate(t=Coalesce(Sum('indirim_tutari'), 0, output_field=DecimalField()))['t']
    toplam_ciro = nakit + kart

    gider_toplam = giderler.aggregate(t=Coalesce(Sum('tutar'), 0, output_field=DecimalField()))['t']
    iptal_toplam = iptallar.aggregate(t=Coalesce(Sum('tutar'), 0, output_field=DecimalField()))['t']

    ikram_listesi = list(
        SiparisItem.objects.filter(adisyon__in=adisyonlar, ikram_mi=True)
        .annotate(deger=ExpressionWrapper(F('adet') * F('urun__fiyat'), output_field=DecimalField()))
        .values('urun__ad')
        .annotate(adet_toplam=Sum('adet'), deger_toplam=Sum('deger'))
        .order_by('-deger_toplam', '-adet_toplam')
    )
    ikram_toplam = sum((item['deger_toplam'] or 0) for item in ikram_listesi)

    en_cok = list(
        SiparisItem.objects.filter(adisyon__in=adisyonlar, ikram_mi=False)
        .annotate(satir_toplami=ExpressionWrapper(F('adet') * F('urun__fiyat'), output_field=DecimalField()))
        .values('urun__ad')
        .annotate(adet_toplam=Sum('adet'), kazanc_toplam=Sum('satir_toplami'))
        .order_by('-kazanc_toplam', '-adet_toplam')
    )

    kategori_raporu = list(
        SiparisItem.objects.filter(adisyon__in=adisyonlar, ikram_mi=False)
        .annotate(satir_toplami=ExpressionWrapper(F('adet') * F('urun__fiyat'), output_field=DecimalField()))
        .values('urun__kategori__ad')
        .annotate(adet_toplam=Sum('adet'), ciro_toplam=Sum('satir_toplami'))
        .order_by('-ciro_toplam')
    )
    for kategori in kategori_raporu:
        kategori['kategori'] = kategori['urun__kategori__ad'] or 'Kategorisiz'

    hazirlanma_raporu = []
    hazirlanma_alanlari_var = (
        _db_column_exists(SiparisItem, 'eklenme_zamani') and
        _db_column_exists(SiparisItem, 'hazir_olma_zamani')
    )
    if hazirlanma_alanlari_var:
        try:
            hazirlanma_sureleri = (
                SiparisItem.objects.filter(
                    adisyon__in=adisyonlar,
                    hazir_mi=True,
                    hazir_olma_zamani__isnull=False,
                    eklenme_zamani__isnull=False,
                )
                .annotate(
                    sure=ExpressionWrapper(
                        F('hazir_olma_zamani') - F('eklenme_zamani'),
                        output_field=fields.DurationField(),
                    )
                )
                .values('urun__ad')
                .annotate(ortalama_sure=Avg('sure'), hazirlanan_adet=Sum('adet'))
            )
            for item in hazirlanma_sureleri:
                if item['ortalama_sure']:
                    dakika = round(item['ortalama_sure'].total_seconds() / 60.0, 1)
                    hazirlanma_raporu.append(
                        {
                            'urun__ad': item['urun__ad'],
                            'hazirlanan_adet': item['hazirlanan_adet'],
                            'ortalama_sure_dk': dakika,
                        }
                    )
        except ProgrammingError:
            hazirlanma_raporu = []
    hazirlanma_raporu = sorted(hazirlanma_raporu, key=lambda row: row['ortalama_sure_dk'], reverse=True)

    garson_raporu = list(
        adisyonlar.values('garson__username')
        .annotate(
            adisyon_sayisi=Count('id'),
            nakit_toplam=Coalesce(Sum('nakit_odenen'), 0, output_field=DecimalField()),
            kart_toplam=Coalesce(Sum('kart_odenen'), 0, output_field=DecimalField()),
        )
        .order_by('-nakit_toplam', '-kart_toplam')
    )
    iptal_sayilari = {
        item['garson__username']: item['iptal_sayisi']
        for item in iptallar.values('garson__username').annotate(iptal_sayisi=Count('id'))
    }
    garson_raporu = [
        {
            'isim': item['garson__username'] or 'Atanmamis',
            'adisyon_sayisi': item['adisyon_sayisi'],
            'ciro': item['nakit_toplam'] + item['kart_toplam'],
            'iptal_sayisi': iptal_sayilari.get(item['garson__username'], 0),
        }
        for item in garson_raporu
        if item['adisyon_sayisi'] > 0
    ]

    saatlik = {}
    for adisyon in adisyonlar:
        saat = timezone.localtime(adisyon.acilis_zamani).hour
        saatlik[saat] = saatlik.get(saat, 0) + float(adisyon.toplam_tutar())
    saatlik_veriler = sorted(saatlik.items())
    zirve_saat = max(saatlik_veriler, key=lambda item: item[1]) if saatlik_veriler else None

    gunluk_kirilim_qs = (
        adisyonlar.annotate(gun=TruncDate('acilis_zamani'))
        .values('gun')
        .annotate(
            nakit_toplam=Coalesce(Sum('nakit_odenen'), 0, output_field=DecimalField()),
            kart_toplam=Coalesce(Sum('kart_odenen'), 0, output_field=DecimalField()),
            adisyon_sayisi=Count('id'),
        )
        .order_by('gun')
    )
    gunluk_kirilim = list(gunluk_kirilim_qs)
    for item in gunluk_kirilim:
        item['ciro'] = item['nakit_toplam'] + item['kart_toplam']

    adisyon_sayisi = adisyonlar.count()
    net_kar = toplam_ciro - gider_toplam
    ortalama_fis = (toplam_ciro / adisyon_sayisi) if adisyon_sayisi else 0
    nakit_oran = round((float(nakit) / float(toplam_ciro) * 100), 1) if toplam_ciro else 0
    kart_oran = round((float(kart) / float(toplam_ciro) * 100), 1) if toplam_ciro else 0

    premium_cards = [
        {
            'label': 'Donem Cirosu',
            'value': f"{format_tl(toplam_ciro)} TL",
            'note': f'{adisyon_sayisi} kapanan hesap',
        },
        {
            'label': 'Net Kar',
            'value': f"{format_tl(net_kar)} TL",
            'note': f'Gider toplam: {format_tl(gider_toplam)} TL',
        },
        {
            'label': 'Iptal Riski',
            'value': f"{format_tl(iptal_toplam)} TL",
            'note': f'{iptallar.count()} iptal kaydi',
        },
        {
            'label': 'Odeme Dagilimi',
            'value': f'Nakit %{nakit_oran} / Kart %{kart_oran}',
            'note': 'Tahsilat kanal dengesi',
        },
    ]

    premium_insights = []
    if en_cok:
        premium_insights.append({
            'title': 'En guclu urun',
            'detail': f"{en_cok[0]['urun__ad']} urunu {en_cok[0]['adet_toplam']} adet ile lider. Gelir katkisi {format_tl(en_cok[0]['kazanc_toplam'])} TL.",
        })
    if zirve_saat:
        premium_insights.append({
            'title': 'Zirve servis saati',
            'detail': f"{zirve_saat[0]}:00 bandi {format_tl(zirve_saat[1])} TL ile gunun en yogun zamani.",
        })
    if hazirlanma_raporu:
        premium_insights.append({
            'title': 'Mutfak darbozazi',
            'detail': f"{hazirlanma_raporu[0]['urun__ad']} ortalama {hazirlanma_raporu[0]['ortalama_sure_dk']} dakika ile en yavas urun.",
        })
    if sube_ozetleri:
        premium_insights.append({
            'title': 'En iyi sube',
            'detail': f"{sube_ozetleri[0]['ad']} subesi {format_tl(sube_ozetleri[0]['ciro'])} TL ciro ile donemin lideri.",
        })
    if garson_raporu:
        premium_insights.append({
            'title': 'En iyi personel',
            'detail': f"{garson_raporu[0]['isim']} {format_tl(garson_raporu[0]['ciro'])} TL ciro ile ekipte onde.",
        })

    return {
        'bugun': bugun,
        'baslangic': baslangic.strftime('%Y-%m-%d'),
        'bitis': bitis.strftime('%Y-%m-%d'),
        'baslangic_date': baslangic,
        'bitis_date': bitis,
        'subeler': subeler,
        'secili_sube': secili_sube,
        'sube_ozetleri': sube_ozetleri,
        'adisyon_sayisi': adisyon_sayisi,
        'gunluk_toplam': format_tl(toplam_ciro),
        'gunluk_toplam_raw': toplam_ciro,
        'nakit_toplam': format_tl(nakit),
        'nakit_toplam_raw': nakit,
        'kart_toplam': format_tl(kart),
        'kart_toplam_raw': kart,
        'indirim_toplam': format_tl(indirim),
        'indirim_toplam_raw': indirim,
        'ikram_toplam': format_tl(ikram_toplam),
        'ikram_toplam_raw': ikram_toplam,
        'gider_toplam': format_tl(gider_toplam),
        'gider_toplam_raw': gider_toplam,
        'iptal_toplam': iptal_toplam,
        'net_kar': format_tl(net_kar),
        'net_kar_raw': net_kar,
        'ortalama_fis': format_tl(ortalama_fis),
        'ortalama_fis_raw': ortalama_fis,
        'en_cok_satanlar': en_cok,
        'kategori_raporu': kategori_raporu,
        'ikram_listesi': ikram_listesi,
        'hazirlanma_raporu': hazirlanma_raporu,
        'garson_raporu': garson_raporu,
        'gider_listesi': giderler,
        'iptal_kayitlari': iptallar,
        'saatlik_veriler': saatlik_veriler,
        'gunluk_kirilim': gunluk_kirilim,
        'premium_cards': premium_cards,
        'premium_insights': premium_insights,
        'zirve_saat': zirve_saat,
        'nakit_oran': nakit_oran,
        'kart_oran': kart_oran,
    }


def _export_excel(context):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Premium Rapor'

    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9F4F1', end_color='D9F4F1', fill_type='solid')

    def add_section(title, rows, headers=None):
        ws.append([])
        ws.append([title])
        ws[f'A{ws.max_row}'].font = bold_font
        ws[f'A{ws.max_row}'].fill = header_fill
        if headers:
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold_font
        for row in rows:
            ws.append(row)

    ws.append(['ELAKI Premium Finans ve Operasyon Raporu'])
    ws.append([f"Tarih araligi: {context['baslangic']} / {context['bitis']}"])
    ws.append([f"Sube filtresi: {context['secili_sube'].ad if context['secili_sube'] else 'Tum Subeler'}"])

    add_section('Mali Ozet', [
        ['Toplam Ciro', float(context['gunluk_toplam_raw'])],
        ['Nakit Tahsilat', float(context['nakit_toplam_raw'])],
        ['Kart Tahsilat', float(context['kart_toplam_raw'])],
        ['Indirimler', float(context['indirim_toplam_raw'])],
        ['Ikram Maliyeti', float(context['ikram_toplam_raw'])],
        ['Toplam Gider', float(context['gider_toplam_raw'])],
        ['Toplam Iptal', float(context['iptal_toplam'])],
        ['Net Kar', float(context['net_kar_raw'])],
        ['Ortalama Fis', float(context['ortalama_fis_raw'])],
    ], ['Kalem', 'Deger'])

    add_section('Premium Ongoruler', [
        [item['title'], item['detail']] for item in context['premium_insights']
    ], ['Baslik', 'Detay'])

    add_section('Urun Performansi', [
        [urun['urun__ad'], urun['adet_toplam'], float(urun['kazanc_toplam'])] for urun in context['en_cok_satanlar']
    ], ['Urun', 'Adet', 'Ciro'])

    add_section('Kategori Performansi', [
        [item['kategori'], item['adet_toplam'], float(item['ciro_toplam'])] for item in context['kategori_raporu']
    ], ['Kategori', 'Adet', 'Ciro'])

    add_section('Sube Karsilastirmasi', [
        [sube['ad'], sube['adisyon_sayisi'], float(sube['nakit']), float(sube['kart']), float(sube['ciro'])] for sube in context['sube_ozetleri']
    ], ['Sube', 'Adisyon', 'Nakit', 'Kart', 'Ciro'])

    add_section('Personel Performansi', [
        [g['isim'], g['adisyon_sayisi'], g['iptal_sayisi'], float(g['ciro'])] for g in context['garson_raporu']
    ], ['Personel', 'Hesap', 'Iptal', 'Ciro'])

    add_section('Gunluk Kirilim', [
        [item['gun'].strftime('%d.%m.%Y'), item['adisyon_sayisi'], float(item['ciro'])] for item in context['gunluk_kirilim']
    ], ['Gun', 'Adisyon', 'Ciro'])

    add_section('Saatlik Yogunluk', [
        [f'{saat}:00', tutar] for saat, tutar in context['saatlik_veriler']
    ], ['Saat', 'Ciro'])

    add_section('Ikramlar', [
        [item['urun__ad'], item['adet_toplam'], float(item['deger_toplam'])] for item in context['ikram_listesi']
    ], ['Urun', 'Adet', 'Deger'])

    add_section('Gider Listesi', [
        [gider.tarih.strftime('%d.%m.%Y'), gider.sube.ad if gider.sube else '-', gider.kategori, gider.aciklama, gider.sorumlu or '-', float(gider.tutar)]
        for gider in context['gider_listesi']
    ], ['Tarih', 'Sube', 'Kategori', 'Aciklama', 'Sorumlu', 'Tutar'])

    add_section('Iptal Kayitlari', [
        [log.zaman.strftime('%d.%m.%Y %H:%M'), log.sube.ad if log.sube else '-', log.urun_adi, log.garson.username if log.garson else '-', log.sebep, float(log.tutar)]
        for log in context['iptal_kayitlari']
    ], ['Zaman', 'Sube', 'Urun', 'Personel', 'Sebep', 'Tutar'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename=ELAKI_Premium_Rapor_{context['baslangic']}_{context['bitis']}.xlsx"
    wb.save(response)
    return response


def _table(data, col_widths=None, header_bg=colors.HexColor('#0E2A47')):
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#dbe4ea')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fbfc')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    return table


def _export_pdf(context):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('ELAKI Premium Finans ve Operasyon Raporu', styles['Title']))
    story.append(Paragraph(
        f"Tarih araligi: {context['baslangic']} - {context['bitis']}<br/>Sube: {context['secili_sube'].ad if context['secili_sube'] else 'Tum Subeler'}",
        styles['BodyText'],
    ))
    story.append(Spacer(1, 12))

    story.append(_table([
        ['Kalem', 'Deger'],
        ['Toplam Ciro', f"{context['gunluk_toplam']} TL"],
        ['Nakit / Kart', f"{context['nakit_toplam']} TL / {context['kart_toplam']} TL"],
        ['Indirim / Ikram', f"{context['indirim_toplam']} TL / {context['ikram_toplam']} TL"],
        ['Gider / Iptal', f"{context['gider_toplam']} TL / {format_tl(context['iptal_toplam'])} TL"],
        ['Net Kar', f"{context['net_kar']} TL"],
        ['Ortalama Fis', f"{context['ortalama_fis']} TL"],
    ], [7 * cm, 10 * cm]))
    story.append(Spacer(1, 12))

    if context['premium_insights']:
        story.append(Paragraph('Premium Ongoruler', styles['Heading2']))
        for item in context['premium_insights']:
            story.append(Paragraph(f"<b>{item['title']}:</b> {item['detail']}", styles['BodyText']))
            story.append(Spacer(1, 4))
        story.append(Spacer(1, 8))

    if context['en_cok_satanlar']:
        story.append(Paragraph('Urun Performansi', styles['Heading2']))
        story.append(_table(
            [['Urun', 'Adet', 'Ciro']] + [[u['urun__ad'], str(u['adet_toplam']), f"{format_tl(u['kazanc_toplam'])} TL"] for u in context['en_cok_satanlar'][:12]],
            [9 * cm, 3 * cm, 5 * cm],
        ))
        story.append(Spacer(1, 10))

    if context['kategori_raporu']:
        story.append(Paragraph('Kategori Performansi', styles['Heading2']))
        story.append(_table(
            [['Kategori', 'Adet', 'Ciro']] + [[k['kategori'], str(k['adet_toplam']), f"{format_tl(k['ciro_toplam'])} TL"] for k in context['kategori_raporu'][:10]],
            [9 * cm, 3 * cm, 5 * cm],
            header_bg=colors.HexColor('#1DA1A1'),
        ))
        story.append(Spacer(1, 10))

    if context['sube_ozetleri']:
        story.append(Paragraph('Sube Karsilastirmasi', styles['Heading2']))
        story.append(_table(
            [['Sube', 'Adisyon', 'Nakit', 'Kart', 'Ciro']] + [
                [s['ad'], str(s['adisyon_sayisi']), format_tl(s['nakit']), format_tl(s['kart']), format_tl(s['ciro'])]
                for s in context['sube_ozetleri'][:10]
            ],
            [5 * cm, 2.2 * cm, 3 * cm, 3 * cm, 3 * cm],
        ))
        story.append(Spacer(1, 10))

    if context['garson_raporu']:
        story.append(Paragraph('Personel Performansi', styles['Heading2']))
        story.append(_table(
            [['Personel', 'Hesap', 'Iptal', 'Ciro']] + [
                [g['isim'], str(g['adisyon_sayisi']), str(g['iptal_sayisi']), f"{format_tl(g['ciro'])} TL"] for g in context['garson_raporu'][:10]
            ],
            [7 * cm, 2.5 * cm, 2.5 * cm, 5 * cm],
            header_bg=colors.HexColor('#2563eb'),
        ))
        story.append(Spacer(1, 10))

    if context['gunluk_kirilim']:
        story.append(Paragraph('Gunluk Kirilim', styles['Heading2']))
        story.append(_table(
            [['Gun', 'Adisyon', 'Ciro']] + [
                [item['gun'].strftime('%d.%m.%Y'), str(item['adisyon_sayisi']), f"{format_tl(item['ciro'])} TL"] for item in context['gunluk_kirilim']
            ],
            [6 * cm, 4 * cm, 6 * cm],
            header_bg=colors.HexColor('#0f766e'),
        ))
        story.append(Spacer(1, 10))

    if context['iptal_kayitlari']:
        story.append(Paragraph('Son Iptal Kayitlari', styles['Heading2']))
        story.append(_table(
            [['Zaman', 'Urun', 'Personel', 'Tutar']] + [
                [log.zaman.strftime('%d.%m.%Y %H:%M'), log.urun_adi, log.garson.username if log.garson else '-', f"{format_tl(log.tutar)} TL"]
                for log in context['iptal_kayitlari'][:12]
            ],
            [5 * cm, 7 * cm, 3.5 * cm, 2.5 * cm],
            header_bg=colors.HexColor('#b91c1c'),
        ))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename=ELAKI_Premium_Rapor_{context['baslangic']}_{context['bitis']}.pdf"
    response.write(pdf)
    return response


@login_required
@user_passes_test(rapor_gorebilir_mi)
def rapor_sayfasi(request, sube_id=None):
    context = _build_report_context(request, sube_id=sube_id)

    if request.GET.get('excel'):
        return _export_excel(context)
    if request.GET.get('pdf'):
        return _export_pdf(context)

    return render(request, 'adisyon_modulu/raporlar.html', context)


@login_required
@user_passes_test(rapor_gorebilir_mi)
def gider_ekle(request):
    if request.method == 'POST':
        sube = Sube.objects.first()

        if not sube:
            return HttpResponse('Sistemde kayitli sube bulunamadi. Lutfen admin panelinden bir sube olusturun.')

        Gider.objects.create(
            sube=sube,
            kategori=request.POST.get('kategori'),
            aciklama=request.POST.get('aciklama'),
            tutar=request.POST.get('tutar'),
            sorumlu=request.POST.get('sorumlu'),
        )
        return redirect('rapor_sayfasi')

    gider_kategoriler = [
        ('Mutfak Alimi', 'Mutfak Alimi'),
        ('Personel Maas/Avans', 'Personel Maas/Avans'),
        ('Kira/Fatura', 'Kira/Fatura'),
        ('Diger', 'Diger'),
    ]

    return render(request, 'adisyon_modulu/gider_ekle.html', {
        'subeler': Sube.objects.all(),
        'kategoriler': gider_kategoriler,
    })
